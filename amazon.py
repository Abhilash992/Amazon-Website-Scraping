# ////////////////LIBRARIES REQUIRED\\\\\\\\\\\\\\\\\
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.common.keys import Keys
import sys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
# ////////////////LIBRARIES REQUIRED\\\\\\\\\\\\\\\\\\\


# ------MAKING A LIST OF LIST Product URL, Product Name, Product Price, Rating, Number of reviews TILL THE END PAGE------
def Extraction(driver, url):
    driver.get(url)
    count = 0
    # Wait for the pagination element to be present
    pagination_strip = driver.find_element(By.CLASS_NAME, 's-pagination-strip')
    No_of_pages = int(pagination_strip.find_element(By.XPATH, './/span[@class="s-pagination-item s-pagination-disabled"]').text)

    extracted = []
    
    for j in range(2, No_of_pages ):
        driver.get(url)
        # Scroll down the page
        driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.END)

        # Find elements using Selenium
        div_class = 'a-section a-spacing-small a-spacing-top-small'
        target_divs = driver.find_elements(By.XPATH, '//div[contains(@class, "a-section a-spacing-small a-spacing-top-small")]')
        for div in target_divs:
            try:
                # Find the <span> element within the current <div>
                span = div.find_element(By.XPATH, './/span[@class="a-size-medium a-color-base a-text-normal"]')
                span_price = div.find_element(By.XPATH, './/span[@class="a-price-whole"]')
                span_rating = div.find_element(By.XPATH, './/span[contains(@aria-label, "out of 5 stars")]')
                span_rating_aria_label = span_rating.get_attribute('aria-label')        
                span_no_rating = div.find_element(By.XPATH, './/span[@class="a-size-base s-underline-text"]')
                # Find the <a> element within the current <div> with a specific class
                link = div.find_element(By.XPATH, './/a[contains(@class, "a-link-normal s-underline-text s-underline-link-text s-link-style a-text-normal")]')

                # Get the text content of the <span> element and the href attribute of the <a> element
                span_text = span.text
                link_href = link.get_attribute('href')
                span_price_text = span_price.text
                rating_start = span_rating_aria_label.index(' out of 5 stars')
                span_rating_text = span_rating_aria_label[:rating_start]    
                span_no_rating_text = span_no_rating.text    

                # Print the span text and the href attribute of the <a> element
                count += 1
                extracted.append([count,link_href,span_text,span_price_text,span_rating_text,span_no_rating_text])
            except:
                # Handle any exceptions that might occur while trying to locate elements
                pass

        # Close the web driver
        desired_label = "Go to page " + str(j)
        url = driver.find_element(By.XPATH, f'//a[@aria-label="{desired_label}"]').get_attribute("href")

    driver.quit()
    return extracted
# ------MAKING A LIST OF LIST Product URL, Product Name, Product Price, Rating, Number of reviews TILL THE END PAGE------



# -------------CREATING AN EXCEL SHEET TO LOAD THE OUTPUT FROM extraction() FUNCTION-------------------------------------

def Create_Excel_File(filename, headers):
    workbook = Workbook()
    sheet = workbook.active

    # Writing column headers
    sheet.append(headers)

    # Calculate and set column widths based on header and content length
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
        
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            
        adjusted_width = (max_length + 2) * 1.2  # Adding some buffer and a factor for conversion
        sheet.column_dimensions[column].width = adjusted_width

    # Save the workbook
    workbook.save(filename)

# -------------CREATING AN EXCEL SHEET TO LOAD THE OUTPUT FROM extraction() FUNCTION-------------------------------------

# -------------LOADING THE OUTPUT FROM extraction() FUNCTION-------------------------------------------------------------

def Loading_Part_1(extracted):

    def write_list_to_excel(filename, sheet_name, start_row, values):
        workbook = load_workbook(filename)
        sheet = workbook[sheet_name]

        # Write values to the specified range
        for j in range (len(values)):
            cell = sheet.cell(row=start_row, column=j + 1)
            cell.value = values[j]

        # Save the workbook
        workbook.save(filename)

    filename = 'Extraction.xlsx'
    sheet_name = 'Sheet'
    for i in range (len(extracted)):
        start_row = i + 2
        write_list_to_excel(filename, sheet_name, start_row, extracted[i])

# -------------LOADING THE OUTPUT FROM extraction() FUNCTION-------------------------------------------------------------


# -------------EXTRACTING DESCRIPTION, PRODUCT DESCRIPTION, ASIN, MANUFACTURER FROM THE LINKS EXTRACTED FROM extraction() FUNCTION-------------------------------------------------------------

def ExtractionAgain(url,i):    
    # Replace with the actual path to your Microsoft Edge WebDriver executable
    edge_service = EdgeService(executable_path="msedgedriver.exe")
    driver = webdriver.Edge(service=edge_service)

    # URL of the webpage you want to scrape
    driver.get(url)
    driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.END)

    try:
        product_description = "Not Available"
        # ------------------------PRODUCT DESCRIPTION-------------------------
        # Find the div element with the specified ID
        div_element = driver.find_element('id', 'productDescription')

        # Find the span element within the div
        span_element = div_element.find_element('tag name', 'span')

        # Extract the text content of the span
        product_description = span_element.text
        # ------------------------PRODUCT DESCRIPTION-------------------------
    except:
        pass
        
        
    try:
        # ------------------------ASIN & MANUFACTURER-------------------------
        Manufacturer = "Not Available"
        ASIN = "Not Available"
            # Find the div element with the specified ID
        div_element = driver.find_element('id', 'detailBullets_feature_div')

        # Find all li elements within the div
        li_elements = div_element.find_elements('tag name', 'li')

        # Extract and print the text content of each li element
        for li_element in li_elements:
            li_text = li_element.text
            if "Manufacturer" in li_text:
                Manufacturer = li_text[15:]
            if "ASIN" in li_text:
                ASIN = li_text[7:]
                break
        # ------------------------ASIN & MANUFACTURER-------------------------
    except:
        pass

    try:
        # ------------------------DESCRIPTION---------------------------------
        description = "Not Available"
        wait = WebDriverWait(driver, 10)
        ul_element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "a-unordered-list.a-vertical.a-spacing-mini")))

        # Find the specific span elements within the desired structure
        span_elements = ul_element.find_elements(By.CSS_SELECTOR, "ul.a-unordered-list.a-vertical.a-spacing-mini > li.a-spacing-mini > span.a-list-item")
        description = ""
        for span_element in span_elements:
            description += "\n" + span_element.text 
        # ------------------------DESCRIPTION---------------------------------
    except:
        pass
    Loading_Part_2(i,description,ASIN,product_description,Manufacturer)

# -------------EXTRACTING DESCRIPTION, PRODUCT DESCRIPTION, ASIN, MANUFACTURER FROM THE LINKS EXTRACTED FROM extraction() FUNCTION-------------------------------------------------------------

# -------------LOADING THE OUTPUT FROM ExtractionAgain() FUNCTION-------------------------------------------------------------

def Loading_Part_2(i,description,ASIN,product_description,Manufacturer):
    
    excel_file_path = 'Extraction.xlsx'
    wb = load_workbook(excel_file_path)
    ws = wb['Sheet']
    if (i == 2):
        ws['G1'] = "DESCRIPTION"
        ws['H1'] = "ASIN"
        ws['I1'] = "PRODUCT DESCRIPTION"
        ws['J1'] = "DESCRIPTION"

    ws['G' +str(i)] = description
    ws['H' +str(i)] = ASIN
    ws['I' +str(i)] = product_description
    ws['J' +str(i)] = Manufacturer
    wb.save(excel_file_path)
    wb.close()

# --------------------------LOADING THE OUTPUT FROM ExtractionAgain() FUNCTION------------------------------------------------

# --------------------------------------TO MODIFY THE EXCEL SHEET-------------------------------------------------------------

def modify_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    # Load the existing Excel workbook
    workbook = openpyxl.load_workbook('Extraction.xlsx')

    # Choose the sheet you want to customize
    sheet = workbook.active

    # Customize the font for the first row
    font = Font(bold=True, size=10)
    for cell in sheet[1]:
        cell.font = font

    # Customize the heading color (pink)
    pink_fill = PatternFill(start_color="ffccd5", end_color="ffccd5", fill_type="solid")
    for cell in sheet[1]:
        cell.fill = pink_fill

    # Customize the alignment for the first row
    for cell in sheet[1]:
        cell.alignment = Alignment(horizontal='left')

    # Customize the background color for remaining rows (blue)
    blue_fill = PatternFill(start_color="a2ded9", end_color="a2ded9", fill_type="solid")
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal='left')

    # Set column widths based on heading lengths
    for col_index, column in enumerate(sheet.columns, start=1):
        heading_length = len(sheet.cell(row=1, column=col_index).value)
        column_letter = openpyxl.utils.get_column_letter(col_index)
        adjusted_width = (heading_length + 2) * 1.2  # Adding some buffer and adjusting units
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Save the changes
    workbook.save('Extraction.xlsx')

# --------------------------------------TO MODIFY THE EXCEL SHEET-------------------------------------------------------------


# Reconfigure standard output to use UTF-8 encoding
sys.stdout.reconfigure(encoding='utf-8')
# Create an EdgeService instance with the WebDriver executable path
edge_service = EdgeService(executable_path="msedgedriver.exe")
# Initialize a WebDriver instance for Edge using the EdgeService.
driver = webdriver.Edge(service=edge_service)

# url of the starting page
url = "https://www.amazon.in/s?k=bags&crid=2M096C61O4MLT&qid=1653308124&sprefix=ba%2Caps%2C283&ref=sr_pg_1"
extracted = Extraction(driver, url)

# file name and headers to create an excel file
filename = 'Extraction.xlsx'
headers = ['SERIAL NO.', 'PRODUCT URL', 'PRODUCT NAME', 'PRODUCT PRICE', 'RATING', 'NUMBER OF REVIEWS']
Create_Excel_File(filename, headers)

# Loading the output of extraction function to Loading_Part_1 function
Loading_Part_1(extracted)


# Loading the already created Extraction.xlsx file to get the maximum number of rows present
wb_obj = openpyxl.load_workbook("Extraction.xlsx")
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row
# Looping all the column 2 (that is the extracted urls) and giving it to ExtractionAgain() function
for i in range(2, max_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=2)
    url = cell_obj.value
    ExtractionAgain(url, i)
driver.quit()




